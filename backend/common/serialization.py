import contextlib
import csv
import json
import re
import urllib.parse
from datetime import date, datetime, time, timedelta
from functools import partial
from io import StringIO
from itertools import chain
from os import PathLike
from pathlib import Path
from typing import Any, IO, Iterable, Literal, TextIO, Union

import charset_normalizer
import openpyxl
import yaml
from babel.numbers import format_decimal
from babel.dates import format_timedelta
from charset_normalizer import from_path
from django.utils.timezone import is_aware, make_naive

fmt_decimal = partial(format_decimal, locale='pt_BR')
fmt_timedelta = partial(format_timedelta, locale='pt_BR')


def dictlist_to_csv(dictlist: list[dict[str, Any]], stream: TextIO, **kwargs):
    """
    Write a list of dictionaries to a CSV format into the provided stream.

    Parameters:
        dictlist (list[dict[str, Any]]): List of dictionaries to convert to CSV.
        stream (TextIO): The output stream to write the CSV data to.

    Returns:
        None

    Raises:
        ValueError: If `dictlist` is empty or contains inconsistent keys.

    Example:
        with open('output.csv', 'w') as f:
            dictlist_to_csv(my_dictlist, f)
    """
    if not dictlist:
        return
    headers = list(dictlist[0].keys())
    headers += sorted(set(chain(*(d.keys() for d in dictlist))) - set(headers))
    dict_writer = csv.DictWriter(stream, headers, **kwargs)
    dict_writer.writeheader()
    dict_writer.writerows(dictlist)


def dictlist_to_xlsxfile(dictlist: list[dict[str, Any]], target: Union[str, Path]):
    lheaders = list(chain(*(d.keys() for d in dictlist)))
    headers = sorted(set(lheaders), key=lambda k: lheaders.index(k))
    colheaders = list(enumerate(headers, 1))
    workbook = openpyxl.Workbook()
    try:
        worksheet = workbook.active
        for icol, header in colheaders:
            worksheet.cell(1, icol, value=header)
        for irow, d in enumerate(dictlist, 2):
            for icol, header in colheaders:
                value = d.get(header)
                if isinstance(value, datetime) and is_aware(value):
                    value = make_naive(value)
                if re.match(r'^(?:19|20)\d{2}-[01]\d-[0123]\d$', str(value)):
                    value = date.fromisoformat(str(value))
                if value is not None and not isinstance(
                    value,
                    (str, int, float, bool, datetime, date, time, timedelta),
                ):
                    value = str(value)
                worksheet.cell(irow, icol, value=value)
        workbook.save(target)
    finally:
        workbook.close()


def dictlist_to_csvstring(dictlist: list[dict[str, Any]], **kwargs) -> str:
    """
    Convert a list of dictionaries to a CSV string.

    Parameters:
        dictlist (list[dict[str, Any]]): List of dictionaries to convert to CSV.

    Returns:
        str: The CSV data as a string.

    Example:
        csv_data = dictlist_to_csvstring(my_dictlist)
    """
    with StringIO() as string_stream:
        dictlist_to_csv(dictlist, string_stream, **kwargs)
        return string_stream.read()


def dictlist_to_csvfile(
    dictlist: list[dict[str, Any]],
    target: Union[str, Path],
    encoding: str = 'utf-8',
    newline: str = '',
):
    """
    Write a list of dictionaries to a CSV file.

    Parameters:
        dictlist (list[dict[str, Any]]): List of dictionaries to convert to CSV.
        target (Union[str, Path]): The file path or Path object where the CSV data should be written.
        encoding (str, optional): The encoding used to write the file. Defaults to 'utf-8'.
        newline (str, optional): The character used to separate a new line. Defaults to ''.

    Returns:
        None

    Raises:
        RuntimeError: If an error occurs during file writing.

    Example:
        dictlist_to_csvfile(my_dictlist, 'output.csv')
    """
    try:
        with Path(target).open('w', encoding=encoding, newline=newline) as stream:
            dictlist_to_csv(dictlist, stream)
    except (OSError, IOError) as e:
        raise RuntimeError(f"Failed to write to file {target}: {e}")


def read_csvstream(stream: TextIO, reader_factory):
    try:
        dialect = csv.Sniffer().sniff(stream.read(4096))
    except csv.Error:
        excel = csv.get_dialect('excel')
        csv.register_dialect(
            'excel-sc',
            delimiter=';',
            quotechar=excel.quotechar,
            escapechar=excel.escapechar,
        )
        dialect = csv.get_dialect('excel-sc')
    stream.seek(0)
    csvreader = reader_factory(stream, dialect=dialect)
    yield from csvreader


def read_csvfile(file: str | Path, reader_factory):
    file = Path(file)
    encoding = charset.encoding if (
        charset := from_path(file).best()) else 'utf-8'
    if (encoding == 'utf-8' or encoding == 'utf_8') and charset and charset.bom:
        encoding = 'utf-8-sig'
    try:
        with file.open('r', encoding=encoding, newline='') as csvstream:
            yield from read_csvstream(csvstream, reader_factory)
    except csv.Error:
        pass


def csvfile_to_dictlist(file: str | Path) -> Iterable[dict[str, Any]]:
    yield from read_csvfile(file, csv.DictReader)


def csvfile_to_lists(file: str | Path) -> Iterable[list[str]]:
    yield from read_csvfile(file, csv.reader)


def csvstring_to_lists(string: str) -> Iterable[list[str]]:
    yield from read_csvstream(StringIO(string), csv.reader)


def to_digits(string):
    return ''.join(c for c in str(string or '') if c.isdigit())


def exclude_empty_str(it: Iterable) -> Iterable[str]:
    for v in it:
        s = v and str(v).strip()
        if s:
            yield s


def format_nup(n):
    n = ''.join(c for c in str(n or '') if c.isdigit()).zfill(17)[-17:]
    return '{0}.{1}/{2}-{3}'.format(n[:-12], n[-12:-6], n[-6:-2], n[-2:])


def format_ndivida(n):
    n = ''.join(c for c in str(n or '') if c.isdigit()).zfill(14)[-14:]
    return '{0}.{1}.{2}/{3}-{4}'.format(n[:-13], n[-13:-10], n[-10:-4], n[-4:-2], n[-2:])


def format_nb_nit(n):
    n = to_digits(str(n))
    n, d = n[:-1], n[-1:]
    lg = []
    while n:
        n, g = n[:-3], n[-3:]
        lg.insert(0, g)
    return '{0}-{1}'.format('.'.join(lg), d)


def format_nu_conta_deposito_judicial(n):
    n = ''.join(c for c in str(n or '') if c.isdigit()).zfill(16)[-16:]
    return '{0}.{1}.{2}-{3}'.format(n[:-12], n[-12:-9], n[-9:-1], n[-1:])


def format_ncnj(n):
    n = ''.join(c for c in str(n or '') if c.isdigit()).zfill(20)[-20:]
    return '{0}-{1}.{2}.{3}.{4}.{5}'.format(n[:-13], n[-13:-11], n[-11:-7], n[-7:-6], n[-6:-4], n[-4:])


def format_cpf(n):
    n = ''.join(c for c in str(n or '') if c.isdigit()).zfill(11)[-11:]
    return '{0}.{1}.{2}-{3}'.format(n[:-8], n[-8:-5], n[-5:-2], n[-2:])


def format_cnpj(n):
    n = ''.join(c for c in str(n or '') if c.isdigit()).zfill(14)[-14:]
    return '{0}.{1}.{2}/{3}-{4}'.format(n[:-12], n[-12:-9], n[-9:-6], n[-6:-2], n[-2:])


def xlsx_to_dictlist(filename: IO[bytes] | str | PathLike[str], worksheet_name: str = None):
    wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
    ws = wb[worksheet_name] if worksheet_name else wb.active
    headers = []
    try:
        for row in ws.iter_rows(values_only=True):
            if not headers:
                headers = [value for value in row]
            else:
                yield dict(zip(headers, row))
    finally:
        wb.close()


def load_json_file(path: str | Path):
    return json.loads(str(charset_normalizer.from_path(Path(path)).best()))


def save_json_file(obj: Any, path: str | Path):
    from django.core.serializers.json import DjangoJSONEncoder
    return Path(path).write_text(json.dumps(obj, cls=DjangoJSONEncoder, indent=4), encoding='utf-8')


def xlsx_to_datatree(filename: IO[bytes] | str | PathLike[str]):
    wb = openpyxl.load_workbook(filename, data_only=True)
    datatree = {}
    try:
        for worksheet in wb.worksheets:
            datatree[worksheet.title] = {
                'data': {}, 'min_col': '_', 'max_col': ' ', 'min_row': 1048577, 'max_row': 0
            }
            for row in worksheet.iter_rows():
                for cell in row:
                    datatree[worksheet.title]['min_col'] = min(
                        cell.column_letter, datatree[worksheet.title]['min_col'],
                    )
                    datatree[worksheet.title]['max_col'] = max(
                        cell.column_letter, datatree[worksheet.title]['max_col'],
                    )
                    datatree[worksheet.title]['min_row'] = min(
                        cell.row, datatree[worksheet.title]['min_row'],
                    )
                    datatree[worksheet.title]['max_row'] = max(
                        cell.row, datatree[worksheet.title]['max_row'],
                    )
                    datatree[worksheet.title]['data'][(
                        cell.column_letter, cell.row
                    )] = cell.value
        return datatree
    finally:
        wb.close()


def tsv_to_nested_list(tsv: str, *, strip_values: bool = False, skip_blanks: bool = False) -> Iterable[list]:
    lines = tsv.splitlines()
    if skip_blanks:
        lines = filter(lambda line: bool(line.strip()), lines)
    if strip_values:
        return [[v.strip() for v in line.split('\t')] for line in lines]
    return [line.split('\t') for line in lines]


class CaseFoldDictReader(csv.DictReader):
    @property
    def fieldnames(self):
        return [s.casefold() for s in super().fieldnames]

    @classmethod
    @contextlib.contextmanager
    def open(cls, file, **kwargs):
        with open(file, mode='r', encoding=kwargs.pop('encoding', None), newline=kwargs.pop('newline', '')) as csvfile:
            yield cls(csvfile, **kwargs)


def csv_parse_iter(string: str):
    dialect = csv.Sniffer().sniff(string)
    yield from csv.reader(StringIO(string), dialect=dialect)


def csv_parse_to_list(string: str):
    return list(csv_parse_iter(string))


def csv_parse_iter_dicts(string: str):
    dialect = csv.Sniffer().sniff(string)
    yield from csv.DictReader(StringIO(string), dialect=dialect)


def csv_parse_to_dictlist(string: str):
    return list(csv_parse_iter_dicts(string))


def ml_yaml_dump(data, *, multiline_style: Literal['|', '>'] = '|', **kwargs):
    def ml_str_representer(dumper: yaml.Dumper, str_data):
        if isinstance(str_data, str) and '\n' in str_data:
            return dumper.represent_scalar('tag:yaml.org,2002:str', str_data, style=multiline_style)
        return dumper.represent_scalar('tag:yaml.org,2002:str', str_data)

    def timedelta_representer(dumper: yaml.Dumper, timedelta_data: timedelta):
        return dumper.represent_scalar('!timedelta', repr(timedelta_data))

    yaml.add_representer(timedelta, timedelta_representer)
    yaml.add_representer(str, ml_str_representer)
    kwargs.setdefault('allow_unicode', True)
    return yaml.dump(data, **kwargs)


class YAML:
    @staticmethod
    def _timedelta_to_iso8601(td: timedelta) -> str:
        total_seconds = td.seconds
        hours, rem = divmod(total_seconds, 3600)
        minutes, secs = divmod(rem, 60)
        result = 'P'
        if td.days:
            result += f'{td.days}D'
        time_part = ''
        if hours:
            time_part += f'{hours}H'
        if minutes:
            time_part += f'{minutes}M'
        if secs or td.microseconds:
            frac = f'{td.microseconds:06d}'.rstrip('0')
            time_part += f'{secs}.{frac}S' if frac else f'{secs}S'
        if time_part:
            result += f'T{time_part}'
        return result or 'PT0S'

    _ISO8601_RE = re.compile(
        r'^P(?:(\d+)D)?(?:T(?:(\d+)H)?(?:(\d+)M)?(?:(\d+(?:\.\d+)?)S)?)?$',
    )

    @classmethod
    def _iso8601_to_timedelta(cls, s: str) -> timedelta:
        m = cls._ISO8601_RE.fullmatch(s)
        if not m:
            raise ValueError(f'Duração ISO 8601 inválida: {s!r}')
        days, hours, minutes, seconds_str = m.groups(default='0')
        seconds = float(seconds_str)
        return timedelta(
            days=int(days),
            hours=int(hours),
            minutes=int(minutes),
            seconds=int(seconds),
            microseconds=round((seconds % 1) * 1_000_000),
        )

    @classmethod
    def _make_dumper(cls, multiline_style: Literal['|', '>']) -> type:
        """Cria um Dumper customizado sem poluir o Dumper global do PyYAML."""

        class CustomDumper(yaml.Dumper):
            allow_unicode = True

        def str_representer(dumper, data):
            if isinstance(data, str) and '\n' in data:
                clean_data = '\n'.join(line.rstrip() for line in data.split('\n'))
                return dumper.represent_scalar('tag:yaml.org,2002:str', clean_data, style=multiline_style)
            return dumper.represent_scalar('tag:yaml.org,2002:str', data)

        def timedelta_representer(dumper, data):
            return dumper.represent_scalar('!duration', cls._timedelta_to_iso8601(data))

        CustomDumper.add_representer(str, str_representer)
        CustomDumper.add_representer(timedelta, timedelta_representer)
        return CustomDumper

    @classmethod
    def _make_loader(cls) -> type:
        """Cria um Loader customizado sem poluir o Loader global do PyYAML."""

        class CustomLoader(yaml.FullLoader):
            pass

        def timedelta_constructor(loader, node):
            return cls._iso8601_to_timedelta(loader.construct_scalar(node))

        CustomLoader.add_constructor('!duration', timedelta_constructor)
        return CustomLoader

    @classmethod
    def dump(cls, data, *, multiline_style: Literal['|', '>'] = '|', **kwargs) -> str:
        kwargs.setdefault('allow_unicode', True)
        kwargs.setdefault('sort_keys', False)
        return yaml.dump(data, Dumper=cls._make_dumper(multiline_style), **kwargs)

    @classmethod
    def load(cls, stream, **kwargs) -> object:
        return yaml.load(stream, Loader=cls._make_loader(), **kwargs)


def normalize_data_url(data_url: str) -> str:
    """
    Normaliza uma Data URL corrigindo o encoding de parâmetros (como espaços e acentos) e removendo atributos
    semanticamente incorretos (como charset para binários). Motivo: propriedade "conteudo" dos componentes digitais do
    Sapiens não obedece às RFCs 2397 e 3986.

    O que essa função faz passo a passo:

    - Separação Segura: Ela quebra a string no exato ponto da primeira vírgula, para não arriscar alterar os dados em
      base64 do seu arquivo.

    - Remoção de Semântica Inválida: Verifica se o MIME type (ex: application/pdf) não começa com text/. Se for esse o
      caso, ela intercepta a chave charset e a descarta automaticamente para limpar a string.

    - URL Encoding Perfeito: Utiliza a função urllib.parse.quote(..., safe="") em todos os valores dos parâmetros
      (como o nome do arquivo). Isso garante que os espaços literais sejam transformados em %20 e o Ó seja transformado
      em %C3%93, tornando a URI 100% aderente às RFCs 2397 e 3986.
    """
    if not data_url.startswith("data:"):
        raise ValueError("A string fornecida não possui o prefixo 'data:'.")

    try:  # Divide a URL em cabeçalho (metadata) e os dados em si (base64)
        header, data_content = data_url.split(",", 1)
    except ValueError:
        raise ValueError("Data URL malformada: vírgula separadora ausente.")

    metadata_string = header[5:]  # Remove o prefixo 'data:' para processar os atributos

    is_base64 = False  # Verifica e isola a flag base64
    if metadata_string.endswith(";base64"):
        is_base64 = True
        metadata_string = metadata_string[:-7]  # Remove o ;base64 do final

    parts = metadata_string.split(";")  # Separa o MIME type dos outros parâmetros
    mime_type = parts[0].strip() if parts[0] else "text/plain"

    normalized_parts = [mime_type]

    is_binary = not mime_type.startswith("text/")  # Identifica se é um arquivo binário para a heurística do charset

    # Processa os atributos adicionais (name=..., charset=...)
    for part in parts[1:]:
        if "=" in part:
            key, value = part.split("=", 1)
            key = key.strip().lower()
            if key == "charset" and is_binary:
                continue
            encoded_value = urllib.parse.quote(value.strip(), safe="")  # safe="" garante que nada passe sem codificação
            normalized_parts.append(f"{key}={encoded_value}")
        else:
            normalized_parts.append(part.strip())

    # Remonta o cabeçalho
    normalized_header = "data:" + ";".join(normalized_parts)
    if is_base64:
        normalized_header += ";base64"

    # Retorna a Data URL completa e normalizada
    return f"{normalized_header},{data_content}"
