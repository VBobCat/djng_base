import warnings
from pathlib import Path
from typing import Dict, List, Literal, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def iter_workbook_data(workbook_path: str | Path):
    workbook = openpyxl.load_workbook(Path(workbook_path), read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                yield worksheet.title, row
    finally:
        workbook.close()


def get_workbook_structure(workbook_path: str | Path):
    workbook = openpyxl.load_workbook(Path(workbook_path), data_only=True)
    retval = {}
    try:
        for worksheet in workbook.worksheets:
            try:
                ws_ref = worksheet.dimensions
            except AttributeError:
                ws_ref = worksheet.calculate_dimension()
            retval[worksheet.title] = {'ref': ws_ref, 'tables': {}}
            try:
                for table_name, table_ref in worksheet.tables.items():
                    retval[worksheet.title]['tables'][table_name] = {'ref': table_ref}
            except AttributeError as ex:
                warnings.warn(str(ex))
                pass
        return retval
    finally:
        workbook.close()


def is_probable_header(row) -> bool:
    return all(cell.value is not None and isinstance(cell.value, str) for cell in row if cell.value != '')


def row_non_empty_cells(row):
    return [cell for cell in row if cell.value not in (None, '')]


def extract_table(sheet: Worksheet, start_row: int, start_col: int, end_col: int) -> List[Dict[str, object]]:
    headers = [sheet.cell(row=start_row, column=col).value for col in range(start_col, end_col + 1)]
    table = []
    row = start_row + 1
    while True:
        values = [sheet.cell(row=row, column=col).value for col in range(start_col, end_col + 1)]
        if all(v in (None, '') for v in values):
            break
        table.append(dict(zip(headers, values)))
        row += 1
    return table


def detect_tables(sheet: Worksheet) -> List[Tuple[str, str, str, List[Dict[str, object]]]]:
    tables = []
    max_row = sheet.max_row
    r = 1
    while r < max_row:
        row = sheet[r]
        if is_probable_header(row):
            next_row = sheet[r + 1] if r + 1 <= max_row else []
            if next_row and len(row_non_empty_cells(row)) == len(row_non_empty_cells(next_row)):
                start_col = next((i + 1 for i, c in enumerate(row) if c.value), 1)
                end_col = max(i + 1 for i, c in enumerate(row) if c.value)
                table_data = extract_table(sheet, r, start_col, end_col)
                start_cell = sheet.cell(row=r, column=start_col).coordinate
                end_cell = sheet.cell(row=r + len(table_data), column=end_col).coordinate
                tables.append((sheet.title, start_cell, end_cell, table_data))
                r += len(table_data) + 1
            else:
                r += 1
        else:
            r += 1
    return tables


def import_all_tabulated_data(file_path: str | Path) -> List[Tuple[str, str, str, List[Dict[str, object]]]]:
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    try:
        all_tables = []
        for sheet in workbook.worksheets:
            sheet_tables = detect_tables(sheet)
            all_tables.extend(sheet_tables)
        return all_tables
    finally:
        workbook.close()


def excel_reader(
        file_path: str | Path,
        *,
        skip_hidden_rows: bool = True,
        data_format: Literal['dict', 'list'] = 'dict',
        single_sheet: int | str = None,
        only_data: bool = False,
):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    try:
        for sheet in workbook.worksheets:
            if isinstance(single_sheet, int) and single_sheet != workbook.worksheets.index(sheet):
                continue
            if isinstance(single_sheet, str) and single_sheet != sheet.title:
                continue
            min_row, max_row = sheet.min_row, sheet.max_row
            irow_header, headers = None, None
            for irow in range(min_row, max_row + 1):
                if skip_hidden_rows and (row := sheet.row_dimensions.get(irow, None)) and row.hidden:
                    continue
                row_values = []
                any_none = False
                all_empty = True
                all_str = True
                for icolumn in range(sheet.min_column, sheet.max_column + 1):
                    cell_value = sheet.cell(row=irow, column=icolumn).value
                    any_none = any_none or cell_value is None
                    all_str = all_str and isinstance(cell_value, str)
                    all_empty = all_empty and cell_value in (None, '')
                    row_values.append(cell_value)
                if all_empty:
                    irow_header, headers = None, None
                    continue
                if all_str and not irow_header:
                    irow_header, headers = irow, row_values
                    if data_format == 'dict':
                        continue
                if data_format == 'dict':
                    if not headers:
                        headers = [f'Col{icolumn+1}' for icolumn in range(sheet.min_column, sheet.max_column + 1)]
                    data = dict(zip(headers, row_values))
                else:
                    data = row_values
                yield data if only_data else {
                    'sheet': sheet.title,
                    'header_row': irow_header,
                    'row': irow,
                    'min_row': min_row,
                    'max_row': max_row,
                    'data': data,
                }
    finally:
        workbook.close()
